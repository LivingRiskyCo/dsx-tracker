"""
Automated Workbook Updater for DSX Opponent Tracker
Fetches latest opponent schedules and updates the Excel workbook
"""

import pandas as pd
from opponent_scraper import OpponentScheduleProcessor
from openpyxl import load_workbook
from datetime import datetime
import os


def update_dsx_workbook(workbook_path, opponent_configs):
    """
    Update the DSX opponent tracker workbook with latest schedules
    
    Args:
        workbook_path: Path to the Excel workbook
        opponent_configs: List of opponent config dicts (name, source, team_id, etc.)
    
    Returns:
        Number of schedules added/updated
    """
    print(f"=== DSX Workbook Auto-Updater ===\n")
    print(f"Workbook: {workbook_path}")
    print(f"Date: {datetime.now().strftime('%Y-%m-%d %H:%M')}\n")
    
    # Step 1: Fetch opponent schedules
    print("Step 1: Fetching opponent schedules...")
    processor = OpponentScheduleProcessor()
    new_schedules = processor.process_opponent_schedules(opponent_configs)
    
    if new_schedules.empty:
        print("  [WARNING] No new schedules fetched")
        return 0
    
    print(f"  [OK] Fetched {len(new_schedules)} schedule entries\n")
    
    # Step 2: Load existing workbook
    print("Step 2: Loading workbook...")
    
    if not os.path.exists(workbook_path):
        print(f"  [ERROR] Workbook not found: {workbook_path}")
        return 0
    
    try:
        wb = load_workbook(workbook_path)
        print(f"  [OK] Loaded workbook\n")
    except Exception as e:
        print(f"  [ERROR] Could not load workbook: {e}")
        return 0
    
    # Step 3: Read existing OppSchedules sheet
    print("Step 3: Reading existing opponent schedules...")
    
    if 'OppSchedules (Paste)' in wb.sheetnames:
        ws = wb['OppSchedules (Paste)']
        
        # Read existing data
        data = []
        headers = []
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i == 0:
                headers = list(row)
            else:
                data.append(row)
        
        existing_df = pd.DataFrame(data, columns=headers) if data else pd.DataFrame()
        print(f"  [OK] Found {len(existing_df)} existing entries\n")
    else:
        existing_df = pd.DataFrame()
        print("  [INFO] 'OppSchedules (Paste)' sheet not found - will create\n")
    
    # Step 4: Merge new schedules with existing (avoid duplicates)
    print("Step 4: Merging schedules...")
    
    if not existing_df.empty:
        # Combine and remove duplicates
        combined_df = pd.concat([existing_df, new_schedules], ignore_index=True)
        combined_df = combined_df.drop_duplicates(
            subset=['OpponentTeam', 'TheirOpponent', 'Date'], 
            keep='last'  # Keep newest version
        )
        new_count = len(combined_df) - len(existing_df)
        print(f"  [OK] Added {new_count} new entries, updated {len(new_schedules) - new_count} existing\n")
    else:
        combined_df = new_schedules
        print(f"  [OK] Added {len(combined_df)} new entries\n")
    
    # Step 5: Write back to workbook
    print("Step 5: Updating workbook...")
    
    try:
        # Remove old sheet if exists
        if 'OppSchedules (Paste)' in wb.sheetnames:
            del wb['OppSchedules (Paste)']
        
        # Create new sheet
        ws = wb.create_sheet('OppSchedules (Paste)')
        
        # Write headers
        for col_num, header in enumerate(combined_df.columns, 1):
            ws.cell(row=1, column=col_num, value=header)
        
        # Write data
        for row_num, row_data in enumerate(combined_df.values, 2):
            for col_num, value in enumerate(row_data, 1):
                ws.cell(row=row_num, column=col_num, value=value)
        
        # Save workbook
        wb.save(workbook_path)
        print(f"  [OK] Workbook updated successfully\n")
        
        return len(combined_df) - len(existing_df)
        
    except Exception as e:
        print(f"  [ERROR] Could not update workbook: {e}")
        return 0


def main():
    """Run the auto-updater"""
    
    # Configuration
    workbook_path = "DSX_U8_2018_Analysis_Fall2025_Dashboard_ImportHelper_v3.xlsx"
    
    # Opponents to track
    opponents = [
        {
            'name': 'BSA Celtic 18B United',
            'source': 'mvysa',
            'team_id': '4264',
            'season': '202509',
            'division': 'MVYSA Fall 2025 B09-3'
        },
        {
            'name': 'BSA Celtic 18B City',
            'source': 'mvysa',
            'team_id': '4263',
            'season': '202509',
            'division': 'MVYSA Fall 2025 B09-3'
        },
        # Add more opponents here as you find their team IDs
        # {
        #     'name': 'Elite FC 2018 Boys Liverpool',
        #     'source': 'gotsport',
        #     'event_id': '45535',
        #     'team_id': 'XXXXX',
        #     'division': 'OCL BU08 Stripes'
        # },
    ]
    
    # Run update
    new_count = update_dsx_workbook(workbook_path, opponents)
    
    # Summary
    print("=" * 50)
    if new_count > 0:
        print(f"SUCCESS: Added {new_count} new opponent schedule(s)")
        print(f"\nNext steps:")
        print(f"  1. Open {workbook_path}")
        print(f"  2. Review 'OppSchedules (Paste)' sheet")
        print(f"  3. Common Opponent Matrix will auto-calculate")
    elif new_count == 0:
        print("INFO: No new schedules to add (all up to date)")
    else:
        print("ERROR: Update failed")
    
    print("=" * 50)
    
    return new_count


if __name__ == "__main__":
    main()

